'''
Use this programme to extract the dates at bicc being reported by projects and check
whether they are correct. This programme produces a workbook with five columns A) project name
B) reported last at bicc date C) a column for manually editing last at bicc dates D) reported next
at bicc date E) column for manually editing next at bicc dates.

If the reported dates needs to be change they should be inserted into columns C and E. The workbook should
then be saved.

The programme highlights in red text if a date has been changed since last quarter (however it doesn't print
out last quarters date; in order to keep the worksheet simple/easy to understand). Red text helps to flag
changes between quarters - so you can check if expected changes have or haven't be reported.

NOTE - the dates are printed in US format. The format of the excel wb needs to be amended
to display UK format dates. This can be easily done and is the best solution for now.

NOTE: if no manual edits are required then the relevant project cell should be left blank. i.e. only enter dates
into the manual edit columns if they need to be changed.

It's important that any manual edits are correctly entered into the workbook - otherwise this may cause issues when
transferring dates back into master and then eventually into the summary dashboard.

To run the programme the correct file paths need to be provided - as highlighted in red below.
'''

from openpyxl import Workbook
from bcompiler.utils import project_data_from_master
from openpyxl.styles import Font

def data_return(project_list, data_key_list, one, two):
    wb = Workbook()
    ws = wb.active

    '''lists project names in ws'''
    for x in range(0, len(project_list)):
        ws.cell(row=x + 2, column=1, value=project_list[x])

    for row_num in range(2, ws.max_row + 1):
        red_text = Font(color="00fc2525")
        project_name = ws.cell(row=row_num, column=1).value
        print(project_name)
        col_start = 2
        if project_name in one:
            for item in data_key_list:
                ws.cell(row=row_num, column=col_start).value = one[project_name][item]
                if one[project_name][item] == None:
                    ws.cell(row=row_num, column=col_start).value = 'Not reported'
                try:
                    if one[project_name][item] != two[project_name][item]:
                        ws.cell(row=row_num, column=col_start).font = red_text
                except KeyError:
                    pass
                col_start += 2
        else:
            ws.cell(row=row_num, column=col_start).value = 'None'
            col_start += 2

    ws.cell(row=1, column=1, value='Project')
    ws.cell(row=1, column=2, value='Last @ BICC')
    ws.cell(row=1, column=3, value='Manual amend: Last @ BICC')
    ws.cell(row=1, column=4, value='Next @ BICC')
    ws.cell(row=1, column=5, value='Manual amend: Next @ BICC')

    return wb


'''1) Specify file paths to master data sources - normally latest and last quarter data'''
one = project_data_from_master(
    'C:\\Users\\Standalone\\Will\\masters folder\\core data\\master_3_2018.xlsx')  # file path to latest quarter
two = project_data_from_master(
    'C:\\Users\\Standalone\\Will\\masters folder\\core data\\master_2_2018.xlsx')  # file path to last quarter

one_list = list(one.keys())
two_list = list(two.keys())
overall_list = sorted(list(set(one_list + two_list)))

data_interest = ['Last time at BICC', 'Next at BICC']

run = data_return(overall_list, data_interest, one, two)

'''2) Specify file path and name of document to be saved'''
run.save(
    'C:\\Users\\Standalone\\Will\\masters folder\\summary_dashboard_docs\\Q3_2018\\test.xlsx')