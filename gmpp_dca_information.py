'''
Programme that compiles a worksheet to show ipa dca ratings from two quarters - last and current.
It takes data from 1) a worksheet currently provided by the ipa - to get the most recent ipa ratings,
and 2) the previous quarters dft master document. It takes ipa dca ratings for projects and places them
into a seperate worksheet highlighting in red text where there has been a change in rating.

The output from this programme is then used to parse any changes in the latest quarters dft master, via the
ipa_dca_to_dft_master programme.

NOTE: ipa dca ratings are always a quarter behind. So the ipa dca ratings will relate to the previous quarter
than what is being reported to BICC.

Things to check:
1) that all projects dca ratings have been recorded. if they haven't there will be gaps in the output document.
Its most likely due to there being naming inconsistencies between ipa and dft. The names need to be consistent
between documents.
'''

from openpyxl import load_workbook, Workbook
from bcompiler.utils import project_data_from_master
from collections import OrderedDict
from openpyxl.utils import column_index_from_string
import datetime
from openpyxl.styles import Font


def put_data_in_dictionary(worksheet):
    d_dict = {}
    for row in worksheet.iter_rows(min_row=2):
        tasks_name = ""
        o = OrderedDict()
        for cell in row:
            if cell.column == 'A':
                tasks_name = cell.value
                d_dict[tasks_name] = o
            else:
                val = worksheet.cell(row=1, column=column_index_from_string(cell.column)).value
                if type(cell.value) == datetime:
                    d_value = datetime(cell.value.year, cell.value.month, cell.value.day)
                    d_dict[tasks_name][val] = d_value
                else:
                    d_dict[tasks_name][val] = cell.value
    try:
        del d_dict[None]
    except KeyError:
        pass

    return d_dict


def put_into_spreadsheet(ipa_data, dft_dict):
    wb = Workbook()
    ws = wb.active
    red_text = Font(color="00fc2525")

    '''places project names in worksheet'''
    project_list = list(ipa_data.keys())
    for i, name in enumerate(project_list):
        ws.cell(row=i + 2, column=1, value=name)

    '''for loop places all ipa dca rating information into worksheet'''
    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=1).value
        if project_name in dft_dict:
            print(project_name)
            ws.cell(row=row_num, column=2).value = dft_dict[project_name]['GMPP - IPA DCA']
        if project_name in ipa_data:
            ws.cell(row=row_num, column=3).value = ipa_data[project_name]['DCA']
            if dft_dict[project_name]['GMPP - IPA DCA'] != ipa_data[project_name]['DCA']:
                ws.cell(row=row_num, column=3).font = red_text

    ws.cell(row=1, column=1, value='Project')
    ws.cell(row=1, column=2, value='IPA rating LAST quarter')
    ws.cell(row=1, column=3, value='IPA rating THIS quarter')

    return wb


'''load GMPP data from IPA excel document, via file path to IPA provided document'''
wb_1 = load_workbook(
    'C:\\Users\\Standalone\\Will\\masters folder\\summary_dashboard_docs\\Q3_2018\\IPA_Q2_1819_DCAs_and_Narratives.xlsx')
data = wb_1.active
ipa_data_1 = put_data_in_dictionary(data)

'''load internal/DfT last quarter data, via file path'''
last_quarter_data = project_data_from_master(
    'C:\\Users\\Standalone\\Will\\masters folder\\core data\\master_2_2018.xlsx')

'''command that runs the programme, specify the file and name of the document being saved'''
run = put_into_spreadsheet(ipa_data_1, last_quarter_data)

run.save(
    'C:\\Users\\Standalone\\Will\\masters folder\\summary_dashboard_docs\\Q3_2018\\test.xlsx')

