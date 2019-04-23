'''
Use this programme to extract the business case stage being reported by projects and check
whether they are correct. This programme produces a workbook with four columns A) project name
B) reported BC stage last quarter C) reported BC stage this quarter, and D) manually edited BC stage.
If the reported BC stage needs to be changed it should be inserted into column D (if no change is required
then leave blank). The workbook should then be saved.

The programme will highlight in red text any changes between the two quarters. This helps to identify
changes between quarters - so you can check if the expected changes have or haven't be reported.

To run the programme the correct file paths need to be provided.
'''

from openpyxl import Workbook
from bcompiler.utils import project_data_from_master
from openpyxl.styles import Font


def data_return(project_list, data_key, list_of_dictionaries):
    wb = Workbook()
    ws = wb.active
    red_text = Font(color="00fc2525")

    '''lists project names in ws'''
    for x in range(0, len(project_list)):
        ws.cell(row=x + 2, column=1, value=project_list[x])

    '''for loop to place bc information into spreadsheet'''
    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=1).value
        print(project_name)
        col_start = 2
        for i, dictionary in enumerate(list_of_dictionaries):
            if project_name in dictionary:
                ws.cell(row=row_num, column=col_start).value = dictionary[project_name][data_key]
                if dictionary[project_name][data_key] == None:
                    ws.cell(row=row_num, column=col_start).value = 'Not reported'
                col_start += 1
                try:
                    if dictionary[project_name][data_key] != dictionary_list[i - 1][project_name][data_key]:
                        ws.cell(row=row_num, column=col_start).font = red_text
                except KeyError:
                    pass
            else:
                ws.cell(row=row_num, column=col_start).value = 'None'
                col_start += 1

    ws.cell(row=1, column=1, value='Project')
    ws.cell(row=1, column=2, value='Last quarter reported BC')
    ws.cell(row=1, column=3, value='This quarter reported BC')
    ws.cell(row=1, column=4, value='Manual edit to BC stage')

    return wb

'''1) Specify path to master data information - usually for latest and last quarter'''
one = project_data_from_master(
    'C:\\Users\\Standalone\\Will\\masters folder\\core data\\master_3_2018.xlsx')
two = project_data_from_master(
    'C:\\Users\\Standalone\\Will\\masters folder\\core data\\master_2_2018.xlsx')

dictionary_list = [two, one]

one_list = list(one.keys())
two_list = list(two.keys())
overall_list = sorted(list(set(one_list + two_list)))

data_interest = 'BICC approval point'

'''running programme'''
run = data_return(overall_list, data_interest, dictionary_list)

'''2) Specify file path and name of document to be saved. I suggest it is {last q info}_bc_stage_from_master'''
run.save(
    'C:\\Users\\Standalone\\Will\\masters folder\\summary_dashboard_docs\\Q3_2018\\test.xlsx')

