'''
This programme collects data from the document created by the 'ipa_dca_info_from_masters' programme, which
should be titled '{quarter_info}_ipa_dca_changes_checker' and inserts changes in ipa dca ratings into the
quarter master data document. Changes to data in the master are highlight in red - so that
they can be easily checked.

To run the programme you need to ensure that you have the correct documents and are providing the correct file
paths to documents as highlighted below. These will need to match the way you have organised your computer
'''

from openpyxl import load_workbook
from collections import OrderedDict
from openpyxl.utils import column_index_from_string
import datetime
from openpyxl.styles import Font

def put_data_in_dictionary(worksheet):
    d_dict = {}
    for row in worksheet.iter_rows(min_row=2):
        tasks_name = ""
        o = OrderedDict()
        for cell in row:
            if cell.column == 'A':
                tasks_name = cell.value
                #print(tasks_name)
                d_dict[tasks_name] = o
            else:
                val = worksheet.cell(row=1, column=column_index_from_string(cell.column)).value
                if type(cell.value) == datetime:
                    d_value = datetime(cell.value.year, cell.value.month, cell.value.day)
                    d_dict[tasks_name][val] = d_value
                else:
                    d_dict[tasks_name][val] = cell.value
    try:
        del d_dict[None]
    except KeyError:
        pass

    return d_dict

def put_into_master(dictionary, worksheet):
    red_text = Font(color="00fc2525")
    for col_num in range(2, worksheet.max_column + 1):
        project_name = ws.cell(row=1, column=col_num).value
        if project_name in dictionary:
            for row_num in range(2, worksheet.max_row + 1):
                if ws.cell(row=row_num, column=1).value == 'GMPP - IPA DCA':
                    ws.cell(row=row_num, column=col_num).value = dictionary[project_name]['IPA rating THIS quarter']
                    if dictionary[project_name]['IPA rating LAST quarter'] == dictionary[project_name]['IPA rating THIS quarter']:
                        print(str(project_name) + ' no change')
                    else:
                        ws.cell(row=row_num, column=col_num).value = dictionary[project_name]['IPA rating THIS quarter']
                        ws.cell(row=row_num, column=col_num).font = red_text
                        print(str(project_name) + ' the IPA DCA change from last quarter and has been recorded in master')

    print('Programme finished')

    return wb

'''1) Specify file path to ipa dca checker document'''
wb_dates = load_workbook('C:\\Users\\Standalone\\Will\\masters folder\\summary_dashboard_docs\\Q3_2018\\'
                         'Q3_1819_ipa_dca_changes_checker.xlsx')
ws_dates = wb_dates.active
data = put_data_in_dictionary(ws_dates)

'''2) Specify file path to master internal spreadsheet'''
wb = load_workbook('C:\\Users\\Standalone\\Will\\masters folder\\master_3_2018.xlsx')
ws = wb.active

amended_master = put_into_master(data, ws)

'''
3) file path to document being save here. Note needs to be same as quarter master data above.
NOTE: this effectively overwrites the master document so make sure you have saved the master before running this 
programme. If you wanted to create a document that doesn't overwrite the master and check changes first, you can type 
a different filename (such as test). However, you will need to save all changes into/overwrite the master at some point
- as it is the solen source of persistent final data.
'''

amended_master.save('C:\\Users\\Standalone\\Will\\masters folder\\master_3_2018.xlsx')

