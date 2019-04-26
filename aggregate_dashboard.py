'''

Programme for creating an aggregate project dashboard

input documents:
1) Dashboard master document - this is an excel file. It should have the dashboard design, with all projects structured
in the correct way (order), but all data fields left blank. Note if project data does not get placed into the correct
part of the master, check that the project name is consistent with the name in master data. The names need to be
exactly the same for information to be released.
2) Master data for two quarters - this will usually be latest and previous quarter

output document:
3) Dashboard with all project data placed into dashboard and formatted correctly.

Instructions:
1) provide path to dashboard master
2) provide path to master data sets
3) change bicc_date variable
4) provide path and specify file name for output document

Supplementary instructions:
These things need to be done to check and assure the data going into the dashboard. Use the other programmes available
for undertaking these tasks.
1) Check that project stage/last at BICC data is correct. This is done via the bc_stage_from_master and
bc_amended_to_master programmes. #TODO check if process needs to be amended for Q1 19/20
2) Check the last at / next at BICC project data is correct. This is done via the bicc_dates_from_master and
bicc_dates_amended_to_master programmes. #TODO check if process needs to be amended for Q1 19/20


Note some manual adjustments may need to be made to:
1) Project WLC totals e.g. Hs2 Phases
2) The last/next at BICC specification. e.g. Hs2 Prog should be changed to 'often'

'''

from openpyxl import load_workbook
from bcompiler.utils import project_data_from_master
import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, IconSet, FormatObject

'''Function that creates dictionary with keys of interest'''
def inital_dict(project_name, data, key_list):
    upper_dictionary = {}
    for name in project_name:
        lower_dictionary = {}

        try:
            p_data = data[name]

            for value in key_list:
                if value in p_data.keys():
                    lower_dictionary[value] = p_data[value]

        except KeyError:
            pass

        upper_dictionary[name] = lower_dictionary

    return upper_dictionary

def all_milestone_data(master_data):
    upper_dict = {}

    for name in master_data:
        p_data = master_data[name]
        lower_dict = {}
        for i in range(1, 50):
            try:
                lower_dict[p_data['Approval MM' + str(i)]] = p_data['Approval MM' + str(i) + ' Forecast / Actual']
            except KeyError:
                lower_dict[p_data['Approval MM' + str(i)]] = p_data['Approval MM' + str(i) + ' Forecast - Actual']

            lower_dict[p_data['Assurance MM' + str(i)]] = p_data['Assurance MM' + str(i) + ' Forecast - Actual']

        for i in range(18, 67):
            lower_dict[p_data['Project MM' + str(i)]] = p_data['Project MM' + str(i) + ' Forecast - Actual']

        upper_dict[name] = lower_dict

    return upper_dict

def add_sop_pend_data(m_data, dict):

    for name in dict:
        try:
            dict[name]['Start of Operation'] = m_data[name]['Start of Operation']
        except KeyError:
            dict[name]['Start of Operation'] = None
        try:
            dict[name]['Project End Date'] = m_data[name]['Project End Date']
        except KeyError:
            dict[name]['Project End Date'] = None

    return dict

'''function for converting dates into concatenated written time periods'''
def concatenate_dates(date):
    today = bicc_date
    if date != None:
        a = (date - today.date()).days
        year = 365
        month = 30
        fortnight = 14
        week = 7
        if a >= 365:
            yrs = int(a / year)
            holding_days_years = a % year
            months = int(holding_days_years / month)
            holding_days_months = a % month
            fortnights = int(holding_days_months / fortnight)
            weeks = int(holding_days_months / week)
        elif 0 <= a <= 365:
            yrs = 0
            months = int(a / month)
            holding_days_months = a % month
            fortnights = int(holding_days_months / fortnight)
            weeks = int(holding_days_months / week)
            # if 0 <= a <=60:
        elif a <= -365:
            yrs = int(a / year)
            holding_days = a % -year
            months = int(holding_days / month)
            holding_days_months = a % -month
            fortnights = int(holding_days_months / fortnight)
            weeks = int(holding_days_months / week)
        elif -365 <= a <= 0:
            yrs = 0
            months = int(a / month)
            holding_days_months = a % -month
            fortnights = int(holding_days_months / fortnight)
            weeks = int(holding_days_months / week)
            # if -60 <= a <= 0:
        else:
            print('something is wrong and needs checking')

        if yrs == 1:
            if months == 1:
                return ('{} yr, {} mth'.format(yrs, months))
            if months > 1:
                return ('{} yr, {} mths'.format(yrs, months))
            else:
                return ('{} yr'.format(yrs))
        elif yrs > 1:
            if months == 1:
                return ('{} yrs, {} mth'.format(yrs, months))
            if months > 1:
                return ('{} yrs, {} mths'.format(yrs, months))
            else:
                return ('{} yrs'.format(yrs))
        elif yrs == 0:
            if a == 0:
                return ('Today')
            elif 1 <= a <= 6:
                return ('This week')
            elif 7 <= a <= 13:
                return ('Next week')
            elif -7 <= a <= -1:
                return ('Last week')
            elif -14 <= a <= -8:
                return ('-2 weeks')
            elif 14 <= a <= 20:
                return ('2 weeks')
            elif 20 <= a <= 60:
                if today.month == date.month:
                    return ('Later this mth')
                elif (date.month - today.month) == 1:
                    return ('Next mth')
                else:
                    return ('2 mths')
            elif -60 <= a <= -15:
                if today.month == date.month:
                    return ('Earlier this mth')
                elif (date.month - today.month) == -1:
                    return ('Last mth')
                else:
                    return ('-2 mths')
            elif months == 12:
                return ('1 yr')
            else:
                return ('{} mths'.format(months))

        elif yrs == -1:
            if months == -1:
                return ('{} yr, {} mth'.format(yrs, -(months)))
            if months < -1:
                return ('{} yr, {} mths'.format(yrs, -(months)))
            else:
                return ('{} yr'.format(yrs))
        elif yrs < -1:
            if months == -1:
                return ('{} yrs, {} mth'.format(yrs, -(months)))
            if months < -1:
                return ('{} yrs, {} mths'.format(yrs, -(months)))
            else:
                return ('{} yrs'.format(yrs))
    else:
        return ('None')

'''function for calculating if confidence has increased decreased'''
def up_or_down(latest_dca, last_dca):

    if latest_dca == last_dca:
        return (int(0))
    elif latest_dca != last_dca:
        if last_dca == 'Green':
            if latest_dca != 'Amber/Green':
                return (int(-1))
        elif last_dca == 'Amber/Green':
            if latest_dca == 'Green':
                return (int(1))
            else:
                return (int(-1))
        elif last_dca == 'Amber':
            if latest_dca == 'Green':
                return (int(1))
            elif latest_dca == 'Amber/Green':
                return (int(1))
            else:
                return (int(-1))
        elif last_dca == 'Amber/Red':
            if latest_dca == 'Red':
                return (int(-1))
            else:
                return (int(1))
        else:
            return (int(1))

'''function for adding concatenated word strings to dictionary.
note probably don't need the above function now, but can tidy up later'''
def final_dict(dict_one, dict_two, con_list, dca_key):
    upper_dict = {}

    for name in dict_one:
        lower_dict = {}
        p_dict_one = dict_one[name]
        for key in p_dict_one:
            if key in con_list:
                try:
                    lower_dict[key] = concatenate_dates(p_dict_one[key])
                except TypeError:
                    lower_dict[key] = 'check data'
            else:
                lower_dict[key] = p_dict_one[key]

        try:
            lower_dict['Change'] = up_or_down(p_dict_one[dca_key], dict_two[name][dca_key])
        except KeyError:
            lower_dict['Change'] = 'NEW'

        upper_dict[name] = lower_dict

    return upper_dict

'''function that places all information into the summary dashboard sheet'''
def placing_excel(dict_one, dict_two):

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        print(project_name)
        if project_name in dict_one:
            ws.cell(row=row_num, column=4).value = dict_one[project_name]['Total Forecast']
            ws.cell(row=row_num, column=6).value = dict_one[project_name]['Change']
            ws.cell(row=row_num, column=7).value = dict_one[project_name]['Departmental DCA']
            ws.cell(row=row_num, column=8).value = dict_one[project_name]['GMPP - IPA DCA last quarter']
            ws.cell(row=row_num, column=9).value = dict_one[project_name]['BICC approval point']
            ws.cell(row=row_num, column=10).value = dict_one[project_name]['Start of Operation']
            ws.cell(row=row_num, column=11).value = dict_one[project_name]['Project End Date']
            ws.cell(row=row_num, column=12).value = dict_one[project_name]['SRO Finance confidence']
            ws.cell(row=row_num, column=13).value = dict_one[project_name]['Last time at BICC']
            ws.cell(row=row_num, column=14).value = dict_one[project_name]['Next at BICC']

    for row_num in range(2, ws.max_row + 1):
        project_name = ws.cell(row=row_num, column=3).value
        if project_name in dict_two:
            ws.cell(row=row_num, column=5).value = dict_two[project_name]['Departmental DCA']

    # Highlight cells that contain RAG text, with background and text the same colour. column E.
    ag_text = Font(color="00a5b700")
    ag_fill = PatternFill(bgColor="00a5b700")
    dxf = DifferentialStyle(font=ag_text, fill=ag_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber/Green", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber/Green",e1)))']
    ws.conditional_formatting.add('e1:e100', rule)

    ar_text = Font(color="00f97b31")
    ar_fill = PatternFill(bgColor="00f97b31")
    dxf = DifferentialStyle(font=ar_text, fill=ar_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber/Red", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber/Red",e1)))']
    ws.conditional_formatting.add('e1:e100', rule)

    red_text = Font(color="00fc2525")
    red_fill = PatternFill(bgColor="00fc2525")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="Red", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Red",E1)))']
    ws.conditional_formatting.add('E1:E100', rule)

    green_text = Font(color="0017960c")
    green_fill = PatternFill(bgColor="0017960c")
    dxf = DifferentialStyle(font=green_text, fill=green_fill)
    rule = Rule(type="containsText", operator="containsText", text="Green", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Green",e1)))']
    ws.conditional_formatting.add('E1:E100', rule)

    amber_text = Font(color="00fce553")
    amber_fill = PatternFill(bgColor="00fce553")
    dxf = DifferentialStyle(font=amber_text, fill=amber_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber",e1)))']
    ws.conditional_formatting.add('e1:e100', rule)

    # Highlight cells that contain RAG text, with background and black text columns G to L.
    ag_text = Font(color="000000")
    ag_fill = PatternFill(bgColor="00a5b700")
    dxf = DifferentialStyle(font=ag_text, fill=ag_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber/Green", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber/Green",G1)))']
    ws.conditional_formatting.add('G1:L100', rule)

    ar_text = Font(color="000000")
    ar_fill = PatternFill(bgColor="00f97b31")
    dxf = DifferentialStyle(font=ar_text, fill=ar_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber/Red", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber/Red",G1)))']
    ws.conditional_formatting.add('G1:L100', rule)

    red_text = Font(color="000000")
    red_fill = PatternFill(bgColor="00fc2525")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="Red", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Red",G1)))']
    ws.conditional_formatting.add('G1:L100', rule)

    green_text = Font(color="000000")
    green_fill = PatternFill(bgColor="0017960c")
    dxf = DifferentialStyle(font=green_text, fill=green_fill)
    rule = Rule(type="containsText", operator="containsText", text="Green", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Green",G1)))']
    ws.conditional_formatting.add('G1:L100', rule)

    amber_text = Font(color="000000")
    amber_fill = PatternFill(bgColor="00fce553")
    dxf = DifferentialStyle(font=amber_text, fill=amber_fill)
    rule = Rule(type="containsText", operator="containsText", text="Amber", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Amber",G1)))']
    ws.conditional_formatting.add('G1:L100', rule)

    # highlighting new projects
    red_text = Font(color="00fc2525")
    white_fill = PatternFill(bgColor="000000")
    dxf = DifferentialStyle(font=red_text, fill=white_fill)
    rule = Rule(type="containsText", operator="containsText", text="NEW", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("NEW",F1)))']
    ws.conditional_formatting.add('F1:F100', rule)

    # assign the icon set to a rule
    first = FormatObject(type='num', val=-1)
    second = FormatObject(type='num', val=0)
    third = FormatObject(type='num', val=1)
    iconset = IconSet(iconSet='3Arrows', cfvo=[first, second, third], percent=None, reverse=None)
    rule = Rule(type='iconSet', iconSet=iconset)
    ws.conditional_formatting.add('F1:F100', rule)

    # change text in last at next at BICC column
    for row_num in range(2, ws.max_row + 1):
        if ws.cell(row=row_num, column=13).value == '-2 weeks':
            ws.cell(row=row_num, column=13).value = 'Last BICC'
        if ws.cell(row=row_num, column=13).value == '2 weeks':
            ws.cell(row=row_num, column=13).value = 'Next BICC'
        if ws.cell(row=row_num, column=13).value == 'Today':
            ws.cell(row=row_num, column=13).value = 'This BICC'
        if ws.cell(row=row_num, column=14).value == '-2 weeks':
            ws.cell(row=row_num, column=14).value = 'Last BICC'
        if ws.cell(row=row_num, column=14).value == '2 weeks':
            ws.cell(row=row_num, column=14).value = 'Next BICC'
        if ws.cell(row=row_num, column=14).value == 'Today':
            ws.cell(row=row_num, column=14).value = 'This BICC'

            # highlight text in bold
    ft = Font(bold=True)
    for row_num in range(2, ws.max_row + 1):
        lis = ['This week', 'Next week', 'Last week', 'Two weeks',
               'Two weeks ago', 'This mth', 'Last mth', 'Next mth',
               '2 mths', '3 mths', '-2 mths', '-3 mths', '-2 weeks',
               'Today', 'Last BICC', 'Next BICC', 'This BICC',
               'Later this mth']
        if ws.cell(row=row_num, column=10).value in lis:
            ws.cell(row=row_num, column=10).font = ft
        if ws.cell(row=row_num, column=11).value in lis:
            ws.cell(row=row_num, column=11).font = ft
        if ws.cell(row=row_num, column=13).value in lis:
            ws.cell(row=row_num, column=13).font = ft
        if ws.cell(row=row_num, column=14).value in lis:
            ws.cell(row=row_num, column=14).font = ft
    return wb


'''keys of interest for current quarter'''
dash_keys = ['Total Forecast', 'Departmental DCA', 'BICC approval point',
            'Project Lifecycle Stage', 'SRO Finance confidence', 'Last time at BICC', 'Next at BICC',
             'GMPP - IPA DCA last quarter']

'''key of interest for previous quarter'''
dash_keys_previous_quarter = ['Departmental DCA']

keys_to_concatenate = ['Start of Operation', 'Project End Date', 'Last time at BICC',
                       'Next at BICC']

'''1) Provide file path to empty dashboard document'''
wb = load_workbook(
    'C:\\Users\\Standalone\\Will\\masters folder\\summary_dashboard_docs\\Q4_2018\\dashboard master_Q4_1819.xlsx')
ws = wb.active

'''2) Provide file path to master data sets'''
data_one = project_data_from_master(
    'C:\\Users\\Standalone\\Will\\masters folder\\core data\\master_4_2018.xlsx')
data_two = project_data_from_master(
    'C:\\Users\\Standalone\\Will\\masters folder\\core data\\master_3_2018.xlsx')

p_names = list(data_one.keys())
#p_names = ['Digital Railway'] # can be useful for checking specific projects/the programme so leaving for now

'''3) Specify data of bicc that is discussing the report. NOTE: Python date format is (YYYY,MM,DD)'''
bicc_date = datetime.datetime(2019, 5, 13)


latest_q_dict = inital_dict(p_names, data_one, dash_keys)
last_q_dict = inital_dict(p_names, data_two, dash_keys_previous_quarter)
m_data = all_milestone_data(data_one)
latest_q_dict_2 = add_sop_pend_data(m_data, latest_q_dict)
merged_dict = final_dict(latest_q_dict_2, last_q_dict, keys_to_concatenate, 'Departmental DCA')
wb = placing_excel(merged_dict, last_q_dict)

'''4) provide file path and specific name of output file.'''
wb.save(
    'C:\\Users\\Standalone\\Will\\masters folder\\summary_dashboard_docs\\Q4_2018\\dashboard_Q4_2018_19.xlsx')