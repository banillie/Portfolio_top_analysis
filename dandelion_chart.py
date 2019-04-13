'''
Programme for planning project wlcs into the dandelion chart. Straight forward programme, however there are a number of
steps that need to done manually, working with excel that are set out elsewhere.

To run the programme you need to specify the file path to master dandelion chart and master data, as well as file
path and name for output file.

Note - at the moment I have not been able to understand and fix the dandelion bubble chart being removed from output
document. This means that the data placed into the output document, needs to be cut and paste into the dandelion master
and then saved as the {quarter info} dandelion chart. No ideal but it is the process for now.

Note - project names need to match those in the master data (other data will not be released). Formulas for adding
totals need to be placed into the master file. Some projects require manual edits to their whole life costs.
'''

from openpyxl import load_workbook
from bcompiler.utils import project_data_from_master

def create_dandelion(wb_name, master_data):
    wb = wb_name
    ws = wb['Table']

    for row_num in range(2, ws.max_row + 1):
        name = ws.cell(row=row_num, column=8).value
        if name in master_data:
            print(name)
            ws.cell(row=row_num, column=9).value = master_data[name]['Total Forecast']

    return wb


'''1) Specify file path to master dandelion document'''
dandelion = load_workbook('C:\\Users\\Standalone\\Will\\masters folder\\dandilion_docs\\Q4_1819_dandelion'
                          '_graph_master.xlsm')

'''2) Specify file path to master data - usually latest quarter'''
data = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\core data\\master_3_2018.xlsx')

wb = create_dandelion(dandelion, data)

'''3) Specify the file path and name for output file'''
wb.save('C:\\Users\\Standalone\\Will\\masters folder\\dandilion_docs\\Q4_1819_dandelion_graph_with_totals.xlsx')

