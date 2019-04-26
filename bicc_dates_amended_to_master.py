'''
This programme collects data from the document created by the 'BICC_dates_from_master' programme, which
should be titled '{quarter_info}_dates_at_bicc_checker' and inserts any manual edits to reported bicc dates
into the quarter master data document.

Changes to data in the master are highlight in red - so that they can be easily checked.

To run the programme you need to ensure that you are providing the correct file paths to documents as highlighted
below.
'''


from openpyxl import load_workbook
from collections import OrderedDict
from openpyxl.utils import column_index_from_string
import datetime
from datetime import datetime
from openpyxl.styles import Font


def put_data_in_dictionary(worksheet):
    d_dict = {}
    for row in worksheet.iter_rows(min_row=2):
        tasks_name = ""
        o = OrderedDict()
        for cell in row:
            if cell.column == 'A':
                tasks_name = cell.value
                print(tasks_name)
                d_dict[tasks_name] = o
            else:
                val = worksheet.cell(row=1, column=column_index_from_string(cell.column)).value
                if type(cell.value) == datetime:
                    d_dict[tasks_name][val] = cell.value
                elif type(cell.value) == str:
                    try:
                        d = cell.value
                        d = datetime.strptime(d, '%d/%m/%Y')
                        d_dict[tasks_name][val] = d
                    except ValueError:
                        d_dict[tasks_name][val] = cell.value
                else:
                    d_dict[tasks_name][val] = cell.value
    try:
        del d_dict[None]
    except KeyError:
        pass

    return d_dict

def put_into_master(dictionary, worksheet):
    red_text = Font(color="00fc2525")
    for col_num in range(2, worksheet.max_column + 1):
        project_name = ws.cell(row=1, column=col_num).value
        if project_name in dictionary:
            if dictionary[project_name]['Manual amend: Last @ BICC'] == None:
                pass
            else:
                for row in range(2, worksheet.max_row + 1):
                    key = ws.cell(row=row, column=1).value
                    if 'Last time at BICC' in key:
                        ws.cell(row=row, column=col_num).value = dictionary[project_name]['Manual amend: Last @ BICC']
                        ws.cell(row=row, column=col_num).font = red_text

    for col_num in range(2, worksheet.max_column + 1):
        project_name = ws.cell(row=1, column=col_num).value
        if project_name in dictionary:
            if dictionary[project_name]['Manual amend: Next @ BICC'] == None:
                pass
            else:
                for row in range(2, worksheet.max_row + 1):
                    key = ws.cell(row=row, column=1).value
                    if 'Next at BICC' in key:
                        ws.cell(row=row, column=col_num).value = dictionary[project_name]['Manual amend: Next @ BICC']
                        ws.cell(row=row, column=col_num).font = red_text
    return wb

'''1) Specify file path to dates at bicc checker document'''
wb_dates = load_workbook('C:\\Users\\Standalone\\Will\\masters folder\\summary_dashboard_docs\\Q4_2018\\'
                         'q4_1819_bicc_dates_check.xlsx')
ws_dates = wb_dates.active
data = put_data_in_dictionary(ws_dates)

'''2) Specify file path to master spreadsheet to be changes'''
wb = load_workbook('C:\\Users\\Standalone\\Will\\masters folder\\core data\\master_4_2018.xlsx')
ws = wb.active

amended_master = put_into_master(data, ws)

'''
3) file path to document being save here. Note needs to be same as quarter master data above.
NOTE: this effectively overwrites the master document so make sure you have saved the master before running this 
programme. If you wanted to create a document that doesn't overwrite the master and check changes first, you can type 
a different filename (such as test). However, you will need to save all changes into/overwrite the master at some point
- as it is the solen source of persistent final data.
'''

amended_master.save('C:\\Users\\Standalone\\Will\\masters folder\\core data\\master_4_2018.xlsx')
